"""
PDOW Program Map -> Excel writer (v2).

Produces a template-based Excel file with the Program Map tab restyled per
user specs:

  Row 1 (banner)
    - Medium-dark blue fill (so navy owl is visible), white bold text
    - Owl image in the far left, Standard Path Order / Course Details /
      Competency Details / <Program> Program Outcomes / Cross Cutting Theme Map
      banners merged across their column groups
    - No border between the owl cell and the banner text to its right
    - Banner text for POs/CCTs wraps to multiple lines

  Row 2 (column headers)
    - Light blue fill, black bold text
    - 3-line-capable row height
    - Columns A-L: Term, Course #, Course, CU, Scope, Ownership, Designation,
      Certificate, Competency Title, Competency Statement, Level, Assessment
    - Columns M+: individual PO names, then CCT names

  Row 3+ (data)
    - Course rows: light gray (F2F2F2), bold
    - Competency rows: white
    - Dark gray thick vertical borders at each section boundary

  Alignment cells (M3+ for rows 3+) have a data-validation dropdown pointing
  to the hidden Menus!A2:A6 sheet (I, R, M, A, X). Free-text still allowed
  (so typing "I, R" for multi-value cases works).

  Tabs kept: Program Map, Program Map Instructions, Standard Path
  Menus tab: preserved but hidden; adds IRMA-X column if not present.
"""
from __future__ import annotations
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image as OpenpyxlImage


# ------------- Palette -------------
BANNER_FILL = PatternFill("solid", start_color="1C3D61", end_color="1C3D61")
BANNER_FONT = Font(bold=True, size=11, color="FFFFFF", name="Arial")

HEADER_FILL = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")
HEADER_FONT = Font(bold=True, size=10, color="000000", name="Arial")

COURSE_FILL = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
WHITE_FILL  = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")

DARK_GRAY = "595959"
LIGHT_GRAY_BORDER = "BFBFBF"
_thin_light = Side(style="thin", color=LIGHT_GRAY_BORDER)
_thick_dark = Side(style="medium", color=DARK_GRAY)
_bottom_banner = Side(style="medium", color=DARK_GRAY)

BODY_BORDER = Border(left=_thin_light, right=_thin_light, top=_thin_light, bottom=_thin_light)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
# LEFT_PADDED: same as LEFT_WRAP but with 1-unit indent (~11px) on the left
# for visual breathing room away from the cell border.
LEFT_PADDED = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)

# Minimum row heights to guarantee ~5pt (7px) of vertical whitespace above and
# below a single line of 10pt text. Excel renders 10pt text at ~13pt line height.
# 13 + 5 + 5 = ~23pt minimum. We use larger values for multi-line content.
MIN_ROW_HEIGHT_SINGLE = 24   # single-line rows
MIN_ROW_HEIGHT_DOUBLE = 38   # two-line wrap
MIN_ROW_HEIGHT_TRIPLE = 52   # three-line wrap
MIN_ROW_HEIGHT_QUAD   = 66   # four-line wrap


# ------------- Column layout -------------
COL_TERM           = 1
COL_COURSE_CODE    = 2
COL_COURSE_NAME    = 3
COL_CU             = 4
COL_SCOPE          = 5
COL_OWNERSHIP      = 6
COL_DESIGNATION    = 7
COL_CERTIFICATE    = 8
COL_COMP_TITLE     = 9
COL_COMP_STMT      = 10
COL_LEVEL          = 11
COL_ASSESSMENT     = 12
FIRST_ALIGNMENT_COL = 13


def _section_boundary_cols(n_po: int) -> set[int]:
    """Columns whose LEFT edge gets the thick dark-gray vertical border."""
    return {
        COL_COURSE_CODE,
        COL_COMP_TITLE,
        FIRST_ALIGNMENT_COL,
        FIRST_ALIGNMENT_COL + n_po,
    }


COL_HEADERS = {
    COL_TERM:          "Term",
    COL_COURSE_CODE:   "Course #",
    COL_COURSE_NAME:   "Course",
    COL_CU:            "CU",
    COL_SCOPE:         "Scope",
    COL_OWNERSHIP:     "Ownership",
    COL_DESIGNATION:   "Designation",
    COL_CERTIFICATE:   "Certificate",
    COL_COMP_TITLE:    "Competency Title",
    COL_COMP_STMT:     "Competency Statement",
    COL_LEVEL:         "Level",
    COL_ASSESSMENT:    "Assessment",
}


BANNER_COURSE   = "Course Details"
BANNER_COMP     = "Competency Details"
BANNER_PO = (
    '{program} Program Outcomes    '
    '"I, R, M" applied to course level to indicate Introduce, Reinforce, and Mastery of Outcome.    '
    '"A" applied to course level to indicate course selected to measure degree-specific outcome.    '
    '"X" indicates outcome coverage at competency level.'
)
BANNER_CCT = (
    'Cross Cutting Theme Map    '
    '"I, R, M" applied to course level to indicate Introduce, Reinforce, and Mastery of Outcome.    '
    'Cross Cutting Themes are not assessed so no "A" will be applied.    '
    '"X" indicates cross cutting theme coverage at competency level.'
)


KEEP_TABS = {"Course Alignment", "Program Map", "Program Map Instructions", "Standard Path"}


def write_program_map(
    model: dict,
    template_path: str | Path,
    output_path: str | Path,
    owl_image_path: str | Path,
    tab_name: str = "Program Map",
    program_code: str | None = None,
) -> Path:
    template_path = Path(template_path)
    output_path = Path(output_path)
    owl_image_path = Path(owl_image_path)

    wb = load_workbook(template_path)
    if tab_name not in wb.sheetnames:
        raise ValueError(f"Tab '{tab_name}' not found. Available: {wb.sheetnames}")

    # 1. Prune tabs
    for sheet_name in list(wb.sheetnames):
        if sheet_name in KEEP_TABS or sheet_name == "Menus":
            continue
        del wb[sheet_name]

    # 2. Rebuild Menus (IRMA-X) and hide it
    _rebuild_menus_tab(wb)

    # 3. Rebuild Program Map
    ws = wb[tab_name]
    pos = model["program_outcomes"]
    ccts = model["ccts"]
    courses = model["courses"]

    label = program_code or _derive_label(model["program"].get("name", ""))

    _reset_sheet(ws)
    _set_column_widths(ws, len(pos), len(ccts))
    _write_banner_row(ws, label, len(pos), len(ccts), owl_image_path)
    _write_header_row(ws, pos, ccts)
    _write_data_rows(ws, courses, pos, ccts)
    _apply_section_borders(ws, courses, len(pos), len(ccts))
    ws.freeze_panes = "J3"

    # 4. Build Course Alignment tab
    ca_ws = _build_course_alignment_tab(wb, model, label, owl_image_path)

    # 5. Rebuild Standard Path tab from our model (template shipped with
    #    BSSCOM-specific course data that we need to replace).
    _build_standard_path_tab(wb, model, label, owl_image_path)

    # 6. Set final tab order:
    #    Standard Path (0) | Program Map Instructions (1) | Program Map (2) |
    #    Course Alignment (3) | Menus (hidden, 4)
    _reorder_tabs(wb, [
        "Standard Path",
        "Program Map Instructions",
        "Program Map",
        "Course Alignment",
        "Menus",
    ])

    # 7. Make Program Map the active landing tab when the file opens.
    # The BSSCOM template carried stale bookViews with firstSheet=5 / activeTab=3
    # that pointed at sheets we deleted, which caused viewers to open on the wrong
    # tab without applying sheet-level freeze panes / seeing the styled content.
    _reset_workbook_view(wb, active_tab_index=wb.sheetnames.index(tab_name))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    # Post-save archive cleanup: openpyxl sometimes re-emits stale
    # commentsDrawing*.vml files inherited from the template. If the shape
    # references inside don't match the fresh comment1.xml, Excel flags
    # the workbook as damaged. Remove the VML file and its rel; openpyxl's
    # written comment1.xml is self-sufficient for cell comments.
    _strip_stale_vml_comments(output_path)

    # Pre-evaluate formula values so viewers that don't evaluate formulas
    # (Claude Excel add-in, Excel Online in certain contexts, some previewers)
    # still show computed values for the Term Total SUMs. Uses LibreOffice
    # headless to recalculate; formulas remain in the file so live editing
    # still works.
    _recalculate_formulas(output_path)

    return output_path


def _recalculate_formulas(output_path: Path):
    """
    Run LibreOffice headless to recalculate all formulas and embed cached
    values next to them. Leaves the formulas intact so downstream edits
    continue to auto-update. Silent no-op if libreoffice isn't available.
    """
    import shutil
    import subprocess
    import tempfile

    if not shutil.which("libreoffice") and not shutil.which("soffice"):
        return  # LO not installed; the formulas will still work in Excel itself

    tmp_dir = tempfile.mkdtemp()
    try:
        binary = shutil.which("libreoffice") or shutil.which("soffice")
        subprocess.run(
            [binary, "--headless", "--calc",
             "--convert-to", "xlsx",
             "--outdir", tmp_dir,
             str(output_path)],
            capture_output=True, timeout=60, check=False,
        )
        # LibreOffice writes the recalculated file with the same basename
        recalc_path = Path(tmp_dir) / output_path.name
        if recalc_path.exists() and recalc_path.stat().st_size > 0:
            shutil.copy(str(recalc_path), str(output_path))
    except (subprocess.TimeoutExpired, OSError):
        pass  # fall through: original file is still valid
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


def _strip_stale_vml_comments(output_path: Path):
    """
    Repack the .xlsx without inherited template cruft that Excel flags as
    "Workbook damaged." Specifically:
      - vmlDrawing*.vml and commentsDrawing*.vml files (legacy comment shapes)
      - comments*.xml files (we don't emit cell comments)
      - All rels entries and Content-Types overrides referencing the above
      - <legacyDrawing .../> elements in sheet XMLs
    """
    import zipfile
    import shutil
    import tempfile
    import re

    src = output_path
    tmp = Path(tempfile.mkstemp(suffix=".xlsx")[1])

    def is_vml(path: str) -> bool:
        return bool(re.search(r"(vmlDrawing|commentsDrawing)\d*\.vml$", path))

    def is_comments(path: str) -> bool:
        return bool(re.search(r"/comments/comment\d+\.xml$|/comments\d+\.xml$", path))

    with zipfile.ZipFile(src, "r") as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.namelist():
            if is_vml(item) or is_comments(item):
                continue

            data = zin.read(item)

            if item.endswith(".rels"):
                text = data.decode()
                # Strip any Relationship pointing at comments or vmlDrawing
                text = re.sub(
                    r'<Relationship\b[^>]*\bType="[^"]*(?:vmlDrawing|comments)[^"]*"[^>]*/>',
                    "",
                    text,
                )
                # Also strip relationships whose Target references those files
                text = re.sub(
                    r'<Relationship\b[^>]*\bTarget="[^"]*(?:vmlDrawing|commentsDrawing|comments/comment\d+\.xml|comments\d+\.xml)[^"]*"[^>]*/>',
                    "",
                    text,
                )
                data = text.encode()

            if item == "[Content_Types].xml":
                text = data.decode()
                # Strip Override entries for vml/comments files
                text = re.sub(
                    r'<Override\b[^>]*PartName="[^"]*(?:vmlDrawing|commentsDrawing|comments/comment\d+\.xml|/comments\d+\.xml)[^"]*"[^>]*/>',
                    "",
                    text,
                )
                data = text.encode()

            if item.startswith("xl/worksheets/sheet") and item.endswith(".xml"):
                text = data.decode()
                # Remove <legacyDrawing .../> — references the VML we removed
                text = re.sub(r'<legacyDrawing\b[^>]*/>', "", text)
                data = text.encode()

            zout.writestr(item, data)

    shutil.move(str(tmp), str(src))


# ============================================================
# Workbook view reset
# ============================================================
def _reset_workbook_view(wb, active_tab_index: int = 0):
    """
    Replace the workbook's bookViews with a single clean view pointing at the
    specified tab. Prevents stale firstSheet/activeTab values from the template
    from making viewers open on the wrong (or hidden) tab.
    """
    from openpyxl.workbook.views import BookView
    view = BookView(
        firstSheet=active_tab_index,
        activeTab=active_tab_index,
    )
    wb.views = [view]


def _reorder_tabs(wb, desired_order: list):
    """
    Reorder worksheets to match desired_order. Tabs named in desired_order that
    exist in the workbook are placed in the given order. Any tabs not named in
    desired_order are appended at the end in their current order.
    """
    present = [name for name in desired_order if name in wb.sheetnames]
    remaining = [name for name in wb.sheetnames if name not in present]
    final_order = present + remaining
    wb._sheets = [wb[name] for name in final_order]


# ============================================================
# Menus tab
# ============================================================
def _rebuild_menus_tab(wb):
    """
    Ensure the Menus tab has an IRMA-X column with I, R, M, A, X values.
    The tab is hidden. Kept for reference even though dropdowns were removed.
    """
    if "Menus" in wb.sheetnames:
        ws = wb["Menus"]
    else:
        ws = wb.create_sheet("Menus")

    target_col = None
    max_scan = max(ws.max_column, 1) + 1
    for c in range(1, max_scan + 1):
        v = ws.cell(row=1, column=c).value
        if v and str(v).strip().upper() in ("IRMA-X", "IRMAX", "IRMA/X"):
            target_col = c
            break
    if target_col is None:
        target_col = 1
        while ws.cell(row=1, column=target_col).value:
            target_col += 1

    ws.cell(row=1, column=target_col).value = "IRMA-X"
    ws.cell(row=1, column=target_col).font = Font(bold=True)
    letters = ["I", "R", "M", "A", "X"]
    for i, letter in enumerate(letters, start=2):
        ws.cell(row=i, column=target_col).value = letter

    for r in range(7, max(7, ws.max_row) + 1):
        ws.cell(row=r, column=target_col).value = None

    ws.sheet_state = "hidden"


# ============================================================
# Derive short program label
# ============================================================
def _derive_label(program_full_name: str) -> str:
    KNOWN = {
        "M.S. Cybersec & Information Assurance": "MSCSIA",
        "B.S. Supply Chain and Operations Management": "BSSCOM",
        "M.S. Nursing Informatics": "MSNI",
        "M.S. Nursing Family Nurse Practitioner": "MSNFNP",
        "M.S. Curriculum and Instruction": "MSCIN",
        "M.S. Curr & Instr - Literacy Specialist": "MSCILS",
        "M.S. Curr & Instr - English Language Learning": "MSCIELL",
    }
    return KNOWN.get(program_full_name, program_full_name)


# ============================================================
# Sheet reset
# ============================================================
def _reset_sheet(ws: Worksheet):
    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))

    max_row = max(ws.max_row, 3)
    max_col = max(ws.max_column, 22)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()
            cell.font = Font()
            cell.alignment = Alignment()
            cell.comment = None  # strip template comments (BSSCOM PO/CCT descriptions)

    ws._images = []
    if hasattr(ws, "data_validations") and ws.data_validations is not None:
        ws.data_validations.dataValidation = []

    # Clear conditional formatting rules inherited from the template — these
    # can override our row fills on columns like Scope/Ownership/Designation.
    if hasattr(ws, "conditional_formatting") and ws.conditional_formatting is not None:
        try:
            ws.conditional_formatting._cf_rules = {}
        except Exception:
            pass

    # Excel "Workbook Repaired" prevention: the BSSCOM template has a
    # legacy VML drawing file (commentsDrawing1.vml) with shape positions
    # for the original comments. openpyxl keeps the file around but
    # regenerates comment1.xml — the mismatch triggers Excel's repair tool.
    # Forcefully reset all comment tracking state on the worksheet.
    if hasattr(ws, "legacy_drawing"):
        ws.legacy_drawing = None
    # Force openpyxl to reset the comments collection
    for row in ws.iter_rows():
        for cell in row:
            if cell._comment is not None:
                cell._comment = None

    # Collapse multiple sheet views to a single default view — the template
    # can come with a second "tabSelected" view that lacks our freeze settings.
    # We want only one view so freeze_panes takes effect when the file is opened.
    if hasattr(ws, "sheet_view"):
        try:
            from openpyxl.worksheet.views import SheetView
            ws.views.sheetView = [SheetView()]
        except Exception:
            pass


# ============================================================
# Column widths
# ============================================================
def _set_column_widths(ws: Worksheet, n_po: int, n_cct: int):
    widths = {
        COL_TERM: 10,
        COL_COURSE_CODE: 10,
        COL_COURSE_NAME: 36,
        COL_CU: 6,
        COL_SCOPE: 11,
        COL_OWNERSHIP: 12,
        COL_DESIGNATION: 16,
        COL_CERTIFICATE: 18,
        COL_COMP_TITLE: 38,
        COL_COMP_STMT: 52,
        COL_LEVEL: 8,
        COL_ASSESSMENT: 14,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    for i in range(n_po + n_cct):
        col = FIRST_ALIGNMENT_COL + i
        ws.column_dimensions[get_column_letter(col)].width = 22


# ============================================================
# Row 1: banner
# ============================================================
def _write_banner_row(ws: Worksheet, program_label: str, n_po: int, n_cct: int,
                      owl_image_path: Path):
    ws.row_dimensions[1].height = 70

    last_col = FIRST_ALIGNMENT_COL + n_po + n_cct - 1
    for c in range(1, last_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = BANNER_FILL
        cell.font = BANNER_FONT
        cell.alignment = CENTER

    # Column A gets no banner text — it holds the owl image only
    ws.cell(row=1, column=COL_COURSE_CODE).value = BANNER_COURSE
    ws.cell(row=1, column=COL_COMP_TITLE).value = BANNER_COMP
    ws.cell(row=1, column=FIRST_ALIGNMENT_COL).value = BANNER_PO.format(program=program_label)
    ws.cell(row=1, column=FIRST_ALIGNMENT_COL + n_po).value = BANNER_CCT

    ws.merge_cells(start_row=1, start_column=COL_COURSE_CODE,
                   end_row=1, end_column=COL_CERTIFICATE)
    ws.merge_cells(start_row=1, start_column=COL_COMP_TITLE,
                   end_row=1, end_column=COL_ASSESSMENT)
    if n_po > 1:
        ws.merge_cells(start_row=1, start_column=FIRST_ALIGNMENT_COL,
                       end_row=1, end_column=FIRST_ALIGNMENT_COL + n_po - 1)
    if n_cct > 1:
        ws.merge_cells(start_row=1, start_column=FIRST_ALIGNMENT_COL + n_po,
                       end_row=1, end_column=FIRST_ALIGNMENT_COL + n_po + n_cct - 1)

    for c in range(1, last_col + 1):
        cell = ws.cell(row=1, column=c)
        existing = cell.border
        cell.border = Border(
            left=existing.left, right=existing.right,
            top=existing.top, bottom=_bottom_banner,
        )

    ws.column_dimensions["A"].width = 10

    if owl_image_path.exists():
        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        from openpyxl.utils.units import pixels_to_EMU

        img = OpenpyxlImage(str(owl_image_path))
        # A1 is approx 75px wide x 93px tall (col width 10, row height 70pt).
        # Keep the owl square and fit it comfortably inside with breathing room.
        target_size_px = 65
        img.width = target_size_px
        img.height = target_size_px

        # Center the owl inside A1
        col_width_px = 75
        row_height_px = 93
        x_offset_px = (col_width_px - target_size_px) // 2
        y_offset_px = (row_height_px - target_size_px) // 2

        marker = AnchorMarker(
            col=0, colOff=pixels_to_EMU(x_offset_px),
            row=0, rowOff=pixels_to_EMU(y_offset_px),
        )
        ext = XDRPositiveSize2D(
            cx=pixels_to_EMU(target_size_px),
            cy=pixels_to_EMU(target_size_px),
        )
        # OneCellAnchor pins from one corner with an explicit extent —
        # no stretching to a second cell, aspect ratio preserved.
        img.anchor = OneCellAnchor(_from=marker, ext=ext)
        ws.add_image(img)


# ============================================================
# Row 2: column headers
# ============================================================
def _write_header_row(ws: Worksheet, pos: list, ccts: list):
    ws.row_dimensions[2].height = 45

    last_col = FIRST_ALIGNMENT_COL + len(pos) + len(ccts) - 1
    for c in range(1, last_col + 1):
        cell = ws.cell(row=2, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BODY_BORDER

    for col, label in COL_HEADERS.items():
        ws.cell(row=2, column=col).value = label

    for i, po in enumerate(pos):
        cell = ws.cell(row=2, column=FIRST_ALIGNMENT_COL + i)
        cell.value = po["name"]
    for i, cct in enumerate(ccts):
        cell = ws.cell(row=2, column=FIRST_ALIGNMENT_COL + len(pos) + i)
        cell.value = cct["name"]


# ============================================================
# Data rows (row 3+)
# ============================================================
def _write_data_rows(ws: Worksheet, courses: list, pos: list, ccts: list):
    row = 3
    for course in courses:
        _write_course_row(ws, row, course, pos, ccts)
        ws.row_dimensions[row].height = 32  # single-line + padding
        row += 1
        for comp in course["competencies"]:
            _write_competency_row(ws, row, comp, pos, ccts)
            ws.row_dimensions[row].height = 44  # 2-3 lines + padding
            row += 1


def _write_course_row(ws: Worksheet, row: int, course: dict, pos: list, ccts: list):
    values = {
        COL_TERM:         course.get("term", ""),
        COL_COURSE_CODE:  course.get("code", ""),
        COL_COURSE_NAME:  course.get("name", ""),
        COL_CU:           course.get("cu") if course.get("cu") is not None else "",
        COL_SCOPE:        course.get("scope", ""),
        COL_OWNERSHIP:    _short_school(course.get("ownership", "")),
        COL_DESIGNATION:  course.get("designation", ""),
        COL_CERTIFICATE:  course.get("certificate", ""),
    }
    for po_idx, po in enumerate(pos):
        values[FIRST_ALIGNMENT_COL + po_idx] = course["po_alignments"].get(po["id"], "")
    for cct_idx, cct in enumerate(ccts):
        values[FIRST_ALIGNMENT_COL + len(pos) + cct_idx] = course["cct_alignments"].get(cct["id"], "")

    _fill_row(ws, row, values, COURSE_FILL, bold=True)


def _write_competency_row(ws: Worksheet, row: int, comp: dict, pos: list, ccts: list):
    level = comp.get("level")
    level_display = int(round(level)) if level is not None else ""
    values = {
        COL_COMP_TITLE:   comp.get("title", ""),
        COL_COMP_STMT:    comp.get("statement", ""),
        COL_LEVEL:        level_display,
        COL_ASSESSMENT:   comp.get("assessment", ""),
    }
    for po_idx, po in enumerate(pos):
        values[FIRST_ALIGNMENT_COL + po_idx] = comp["po_alignments"].get(po["id"], "")
    for cct_idx, cct in enumerate(ccts):
        values[FIRST_ALIGNMENT_COL + len(pos) + cct_idx] = comp["cct_alignments"].get(cct["id"], "")

    _fill_row(ws, row, values, WHITE_FILL, bold=False)


def _fill_row(ws: Worksheet, row: int, values: dict, fill: PatternFill, bold: bool):
    total_cols = max(ws.max_column, max(values.keys(), default=1))
    for col in range(1, total_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.border = BODY_BORDER
        cell.font = Font(bold=bold, size=10, name="Arial")
        # Always set alignment so empty cells in the row also get wrap + vertical center
        if col in (COL_COURSE_NAME, COL_COMP_TITLE, COL_COMP_STMT):
            cell.alignment = LEFT_PADDED
        else:
            cell.alignment = CENTER
        # Set value only when provided
        if col in values:
            v = values[col]
            cell.value = v if v != "" else None


def _short_school(school_name: str) -> str:
    return {
        "School of Business": "WSB",
        "School of Technology": "WST",
        "School of Health Leavitt": "WSH",
        "General Education": "Gen Ed",
    }.get(school_name, school_name)


# ============================================================
# Section borders
# ============================================================
def _apply_section_borders(ws: Worksheet, courses: list, n_po: int, n_cct: int):
    boundaries = _section_boundary_cols(n_po)
    total_data_rows = sum(1 + len(c["competencies"]) for c in courses)
    last_row = 2 + total_data_rows

    for row in range(2, last_row + 1):
        for col in boundaries:
            cell = ws.cell(row=row, column=col)
            existing = cell.border
            cell.border = Border(
                left=_thick_dark,
                right=existing.right,
                top=existing.top,
                bottom=existing.bottom,
            )
            prev = ws.cell(row=row, column=col - 1)
            prev_ex = prev.border
            prev.border = Border(
                left=prev_ex.left,
                right=_thick_dark,
                top=prev_ex.top,
                bottom=prev_ex.bottom,
            )


# ============================================================
# Course Alignment tab — transposed course-level view
# ============================================================
# Layout:
#   Row 1: Navy banner with owl (A1) + "Course Alignment" title section +
#          "Courses" banner spanning all course columns
#   Row 2: Light-blue column headers: col A = "Outcome / Theme",
#          col B = "Description", cols C+ = course headers (two-line: "E123"
#          over "Cybersecurity Fundamentals")
#   Row 3: Navy section banner "Program Outcomes" spanning the full width
#   Rows 4..4+n_po-1: one row per PO, with IRMA-X values in each course column
#   Row 4+n_po: Navy section banner "Cross Cutting Themes" spanning the full width
#   Following rows: one per CCT
# Freeze panes: C3 (rows 1-2 frozen; cols A-B frozen)

CA_TAB_NAME = "Course Alignment"
CA_COL_NAME = 1     # A: Outcome / Theme name
CA_COL_DESC = 2     # B: Description
CA_FIRST_COURSE_COL = 3  # C: first course column


def _build_course_alignment_tab(wb, model: dict, program_label: str,
                                  owl_image_path: Path) -> Worksheet:
    """Create and populate the Course Alignment tab. Returns the worksheet."""
    if CA_TAB_NAME in wb.sheetnames:
        del wb[CA_TAB_NAME]
    ws = wb.create_sheet(CA_TAB_NAME)

    pos = model["program_outcomes"]
    ccts = model["ccts"]
    courses = model["courses"]
    n_courses = len(courses)
    last_col = CA_FIRST_COURSE_COL + n_courses - 1

    # ---- Column widths ----
    ws.column_dimensions["A"].width = 34  # Outcome / Theme
    ws.column_dimensions["B"].width = 60  # Description
    for i in range(n_courses):
        ws.column_dimensions[get_column_letter(CA_FIRST_COURSE_COL + i)].width = 16

    # ---- Row 1: Navy banner ----
    ws.row_dimensions[1].height = 70
    for c in range(1, last_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = BANNER_FILL
        cell.font = BANNER_FONT
        cell.alignment = CENTER

    # Col A: holds the owl; no text
    # Col B: "Course Alignment" title
    ws.cell(row=1, column=CA_COL_DESC).value = (
        f"{program_label} Course Alignment Map    "
        f'"I, R, M" applied to indicate Introduce, Reinforce, and Mastery of Outcome.    '
        f'"A" applied to indicate course selected to measure degree-specific outcome.'
    )
    if n_courses > 0:
        ws.cell(row=1, column=CA_FIRST_COURSE_COL).value = "Courses"
        if n_courses > 1:
            ws.merge_cells(
                start_row=1, start_column=CA_FIRST_COURSE_COL,
                end_row=1, end_column=last_col,
            )

    # Bottom banner border
    for c in range(1, last_col + 1):
        cell = ws.cell(row=1, column=c)
        existing = cell.border
        cell.border = Border(
            left=existing.left, right=existing.right,
            top=existing.top, bottom=_bottom_banner,
        )

    # Insert owl in A1
    if owl_image_path.exists():
        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        from openpyxl.utils.units import pixels_to_EMU

        img = OpenpyxlImage(str(owl_image_path))
        target_size_px = 65
        img.width = target_size_px
        img.height = target_size_px

        col_a_width_px = 238  # col width 34 -> ~238px
        row_height_px = 93
        x_offset_px = (col_a_width_px - target_size_px) // 2
        y_offset_px = (row_height_px - target_size_px) // 2

        marker = AnchorMarker(
            col=0, colOff=pixels_to_EMU(x_offset_px),
            row=0, rowOff=pixels_to_EMU(y_offset_px),
        )
        ext = XDRPositiveSize2D(
            cx=pixels_to_EMU(target_size_px),
            cy=pixels_to_EMU(target_size_px),
        )
        img.anchor = OneCellAnchor(_from=marker, ext=ext)
        ws.add_image(img)

    # ---- Row 2: Column headers (light blue) ----
    ws.row_dimensions[2].height = 55  # tall enough for two-line course headers
    for c in range(1, last_col + 1):
        cell = ws.cell(row=2, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BODY_BORDER

    ws.cell(row=2, column=CA_COL_NAME).value = "Outcome / Theme"
    ws.cell(row=2, column=CA_COL_DESC).value = "Description"

    # Course headers: "E123\nCybersecurity Fundamentals"
    for i, course in enumerate(courses):
        col = CA_FIRST_COURSE_COL + i
        code = course.get("code") or ""
        name = course.get("name") or ""
        cell = ws.cell(row=2, column=col)
        cell.value = f"{code}\n{name}" if code and name else (code or name)

    # ---- Rows 3+: sections ----
    row = 3
    row = _write_ca_section_banner(ws, row, "Program Outcomes", last_col)
    for po in pos:
        _write_ca_outcome_row(ws, row, po, courses, "po")
        row += 1

    row = _write_ca_section_banner(ws, row, "Cross Cutting Themes", last_col)
    for cct in ccts:
        _write_ca_outcome_row(ws, row, cct, courses, "cct")
        row += 1

    # ---- Freeze panes: rows 1-2 + cols A-B frozen ----
    ws.freeze_panes = "C3"

    return ws


def _write_ca_section_banner(ws: Worksheet, row: int, label: str, last_col: int) -> int:
    """Write a full-width navy banner row (PO section or CCT section divider). Returns next row index."""
    ws.row_dimensions[row].height = 24
    for c in range(1, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = BANNER_FILL
        cell.font = Font(bold=True, size=11, color="FFFFFF", name="Arial")
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.cell(row=row, column=1).value = label
    if last_col > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    return row + 1


def _write_ca_outcome_row(ws: Worksheet, row: int, outcome: dict, courses: list,
                           kind: str):
    """
    Write one row for a PO or CCT:
      col A: outcome/theme name
      col B: description (wrapped)
      col C+: IRMA-X values for this outcome × each course
    kind: "po" uses course["po_alignments"], "cct" uses course["cct_alignments"]
    """
    # Name cell
    name_cell = ws.cell(row=row, column=CA_COL_NAME)
    name_cell.value = outcome.get("name", "")
    name_cell.fill = WHITE_FILL
    name_cell.font = Font(bold=True, size=10, name="Arial")
    name_cell.alignment = LEFT_PADDED
    name_cell.border = BODY_BORDER

    # Description cell
    description = outcome.get("description", "") or ""
    desc_cell = ws.cell(row=row, column=CA_COL_DESC)
    desc_cell.value = description
    desc_cell.fill = WHITE_FILL
    desc_cell.font = Font(size=10, name="Arial")
    desc_cell.alignment = LEFT_PADDED
    desc_cell.border = BODY_BORDER

    # Dynamic row height: ~60 chars per line at col B width 60 chars.
    # Compute line estimate, then add ~5pt above/below for vertical padding.
    name = outcome.get("name", "") or ""
    name_lines = max(1, (len(name) + 32) // 33)  # col A is ~34 chars wide
    desc_lines = max(1, (len(description) + 59) // 60)  # col B is ~60 chars wide
    lines = max(name_lines, desc_lines)
    # 13pt per line of text + 10pt of padding (~5pt top + ~5pt bottom)
    ws.row_dimensions[row].height = max(MIN_ROW_HEIGHT_SINGLE, 13 * lines + 10)

    # One cell per course with the IRMA-X letters for this outcome
    align_key = "po_alignments" if kind == "po" else "cct_alignments"
    outcome_id = outcome["id"]
    for i, course in enumerate(courses):
        col = CA_FIRST_COURSE_COL + i
        cell = ws.cell(row=row, column=col)
        cell.value = course.get(align_key, {}).get(outcome_id, "")
        cell.fill = WHITE_FILL
        cell.font = Font(size=10, name="Arial")
        cell.alignment = CENTER
        cell.border = BODY_BORDER


# ============================================================
# Standard Path tab — single-column course pathway view
# ============================================================
# Layout:
#   Row 1: Navy banner with owl (A1) + "<Program> Proposed Standard Path" title
#   Row 2: Light-blue column headers — Term, Scope, Course Code, Course Name,
#          CU, Assessment, Course Scope/Description, Competency Count
#   Rows 3+: Per-term groups:
#     - one row per course (sorted by std_path_order within term)
#     - one "Term X Total" row with SUM formula for CU
# Freeze panes: A3

SP_TAB_NAME = "Standard Path"
SP_COL_STD_PATH    = 1   # A: Standard Path Order
SP_COL_TERM        = 2   # B
SP_COL_SCOPE       = 3   # C
SP_COL_CODE        = 4   # D
SP_COL_NAME        = 5   # E
SP_COL_CU          = 6   # F
SP_COL_COMP_COUNT  = 7   # G (moved before Assessment)
SP_COL_ASSESSMENT  = 8   # H
SP_COL_DESCRIPTION = 9   # I
SP_LAST_COL = 9


def _build_standard_path_tab(wb, model: dict, program_label: str,
                              owl_image_path: Path) -> Worksheet:
    """Rebuild the Standard Path tab from the model. Returns the worksheet."""
    if SP_TAB_NAME in wb.sheetnames:
        ws = wb[SP_TAB_NAME]
        _reset_sheet(ws)
    else:
        ws = wb.create_sheet(SP_TAB_NAME)

    courses = model["courses"]

    # ---- Column widths ----
    widths = {
        SP_COL_STD_PATH:    8,
        SP_COL_TERM:        6,
        SP_COL_SCOPE:       12,
        SP_COL_CODE:        10,
        SP_COL_NAME:        38,
        SP_COL_CU:          5,
        SP_COL_COMP_COUNT:  8,
        SP_COL_ASSESSMENT:  13,
        SP_COL_DESCRIPTION: 75,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # ---- Row 1: Navy banner ----
    ws.row_dimensions[1].height = 70
    for c in range(1, SP_LAST_COL + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = BANNER_FILL
        cell.font = BANNER_FONT
        cell.alignment = CENTER
        existing = cell.border
        cell.border = Border(
            left=existing.left, right=existing.right,
            top=existing.top, bottom=_bottom_banner,
        )
    ws.cell(row=1, column=SP_COL_TERM).value = f"{program_label} Proposed Standard Path"
    ws.merge_cells(
        start_row=1, start_column=SP_COL_TERM,
        end_row=1, end_column=SP_LAST_COL,
    )

    # Owl in A1
    if owl_image_path.exists():
        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        from openpyxl.utils.units import pixels_to_EMU

        img = OpenpyxlImage(str(owl_image_path))
        target_size_px = 60
        img.width = target_size_px
        img.height = target_size_px

        col_a_width_px = int(widths[SP_COL_TERM] * 7)  # approx chars→px
        row_height_px = 93
        x_offset_px = max(0, (col_a_width_px - target_size_px) // 2)
        y_offset_px = (row_height_px - target_size_px) // 2

        marker = AnchorMarker(
            col=0, colOff=pixels_to_EMU(x_offset_px),
            row=0, rowOff=pixels_to_EMU(y_offset_px),
        )
        ext = XDRPositiveSize2D(
            cx=pixels_to_EMU(target_size_px),
            cy=pixels_to_EMU(target_size_px),
        )
        img.anchor = OneCellAnchor(_from=marker, ext=ext)
        ws.add_image(img)

    # ---- Row 2: Column headers ----
    ws.row_dimensions[2].height = 54  # fits up to 3 lines of wrapped 10pt text
    headers = {
        SP_COL_STD_PATH:    "Standard Path Order",
        SP_COL_TERM:        "Term",
        SP_COL_SCOPE:       "Scope",
        SP_COL_CODE:        "Course Code",
        SP_COL_NAME:        "Course Name",
        SP_COL_CU:          "CU",
        SP_COL_COMP_COUNT:  "Comp Count",
        SP_COL_ASSESSMENT:  "Assessment",
        SP_COL_DESCRIPTION: "Course Scope / Description",
    }
    for c in range(1, SP_LAST_COL + 1):
        cell = ws.cell(row=2, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BODY_BORDER
    for col, label in headers.items():
        ws.cell(row=2, column=col).value = label

    # ---- Rows 3+: Per-term course groups + term totals ----
    # Group courses by term while preserving sort order (already sorted by extractor)
    from collections import OrderedDict
    by_term: "OrderedDict[str, list]" = OrderedDict()
    for course in courses:
        term = str(course.get("term") or "?")
        by_term.setdefault(term, []).append(course)

    row = 3
    for term, term_courses in by_term.items():
        term_first_row = row
        for course in term_courses:
            _write_sp_course_row(ws, row, course)
            row += 1
        term_last_row = row - 1
        _write_sp_term_total_row(ws, row, term, term_first_row, term_last_row)
        ws.row_dimensions[row].height = 24
        row += 1

    # ---- Freeze panes: rows 1-2 + cols A-D (through Course Code) frozen ----
    ws.freeze_panes = "E3"
    return ws


def _write_sp_course_row(ws: Worksheet, row: int, course: dict):
    """One course row on the Standard Path tab."""
    values = {
        SP_COL_STD_PATH:    course.get("std_path_order") if course.get("std_path_order") is not None else "",
        SP_COL_TERM:        course.get("term", ""),
        SP_COL_SCOPE:       course.get("scope", ""),
        SP_COL_CODE:        course.get("code", ""),
        SP_COL_NAME:        course.get("name", ""),
        SP_COL_CU:          course.get("cu") if course.get("cu") is not None else "",
        SP_COL_COMP_COUNT:  len(course.get("competencies", [])),
        SP_COL_ASSESSMENT:  _primary_assessment(course),
        SP_COL_DESCRIPTION: _course_description_summary(course),
    }
    # Row height: give description plenty of breathing room.
    # 72pt fits ~4 lines of 11pt text with clear white space above and below.
    ws.row_dimensions[row].height = 72

    for col in range(1, SP_LAST_COL + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = WHITE_FILL
        cell.border = BODY_BORDER
        cell.font = Font(size=10, name="Arial")
        v = values.get(col, "")
        cell.value = v if v != "" else None
        if col in (SP_COL_NAME, SP_COL_DESCRIPTION):
            cell.alignment = LEFT_PADDED
        else:
            cell.alignment = CENTER


def _write_sp_term_total_row(ws: Worksheet, row: int, term: str,
                               first_course_row: int, last_course_row: int):
    """Light-gray summary row with SUM formula for the term's CU total."""
    for col in range(1, SP_LAST_COL + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = COURSE_FILL
        cell.border = BODY_BORDER
        cell.font = Font(bold=True, size=10, name="Arial")
        cell.alignment = CENTER

    ws.cell(row=row, column=SP_COL_NAME).value = f"Term {term} Total"
    ws.cell(row=row, column=SP_COL_NAME).alignment = Alignment(
        horizontal="right", vertical="center"
    )
    # CU SUM formula
    cu_letter = get_column_letter(SP_COL_CU)
    ws.cell(row=row, column=SP_COL_CU).value = (
        f"=SUM({cu_letter}{first_course_row}:{cu_letter}{last_course_row})"
    )
    # Competency count SUM formula
    cc_letter = get_column_letter(SP_COL_COMP_COUNT)
    ws.cell(row=row, column=SP_COL_COMP_COUNT).value = (
        f"=SUM({cc_letter}{first_course_row}:{cc_letter}{last_course_row})"
    )


def _primary_assessment(course: dict) -> str:
    """
    Courses don't carry a direct assessment field in our model — the assessment
    lives on each competency. Return the most common competency assessment
    for the course, or blank if none.
    """
    from collections import Counter
    comps = course.get("competencies", [])
    assessments = [c.get("assessment", "") for c in comps if c.get("assessment")]
    if not assessments:
        return ""
    counter = Counter(assessments)
    if len(counter) == 1:
        return assessments[0]
    most_common, _ = counter.most_common(1)[0]
    return most_common


def _course_description_summary(course: dict, max_chars: int = 450) -> str:
    """
    Pick the right content field based on the course's scope, then summarize
    it to fit in roughly 4 lines at the current column width (~75 chars per
    line → ~450 chars total, allowing for wrap + padding).

    Scope rules (flexible):
      - New / Redesign / Redev: prefer scope_notes; fall back to description
        if scope_notes is blank.
      - Existing: prefer description; fall back to scope_notes if description
        is blank.
      - Otherwise: use whichever is non-empty.

    Summarization: keep whole sentences. Truncate at the last sentence break
    that fits, appending an ellipsis if content was cut. If no sentence
    boundary exists in range, hard-cut at a word boundary with ellipsis.
    """
    scope = (course.get("scope") or "").strip().lower()
    scope_notes = (course.get("scope_notes") or "").strip()
    description = (course.get("description") or "").strip()

    # Strip any "COURSE LEVEL SCOPE NOTES:" prefix from scope_notes
    import re as _re
    scope_notes = _re.sub(r"^\s*COURSE LEVEL SCOPE NOTES\s*:\s*", "", scope_notes,
                          flags=_re.IGNORECASE)

    # Pick preferred field by scope, fall back to the other
    if scope in ("new", "redesign", "redev"):
        text = scope_notes or description
    elif scope == "existing":
        text = description or scope_notes
    else:
        text = description or scope_notes

    if not text:
        return ""

    # Normalize whitespace
    text = _re.sub(r"\s+", " ", text).strip()

    if len(text) <= max_chars:
        return text

    # Try to cut at the last sentence boundary within max_chars
    window = text[:max_chars]
    # Look for end-of-sentence punctuation followed by a space
    match = list(_re.finditer(r"[.!?]\s", window))
    if match:
        last = match[-1]
        cut = last.end()
        return text[:cut].rstrip() + " …"

    # No sentence boundary — hard-cut at a word boundary
    last_space = window.rfind(" ")
    if last_space > max_chars * 0.5:
        return text[:last_space].rstrip() + " …"
    return window.rstrip() + " …"


# ============================================================
# Standalone test
# ============================================================
if __name__ == "__main__":
    import json
    import sys

    here = Path(__file__).parent
    model_path = here / "data" / "mscsia_model.json"
    template = here / "template_BSSCOM.xlsx"
    owl = here / "owl.png"
    output = here / "data" / "MSCSIA_Program_Map_v2.xlsx"

    if not model_path.exists():
        print(f"No model at {model_path}", file=sys.stderr)
        sys.exit(1)

    with model_path.open() as f:
        model = json.load(f)

    # Inject sample alignment values to exercise the dropdowns and multi-value cells
    if model["courses"] and model["program_outcomes"]:
        c0 = model["courses"][0]
        po_id = model["program_outcomes"][0]["id"]
        c0["po_alignments"][po_id] = "I, R"
        if len(model["program_outcomes"]) > 1:
            c0["po_alignments"][model["program_outcomes"][1]["id"]] = "I"
        if model["ccts"]:
            c0["cct_alignments"][model["ccts"][0]["id"]] = "M"
        if c0["competencies"]:
            c0["competencies"][0]["po_alignments"][po_id] = "X"

    result = write_program_map(
        model=model,
        template_path=template,
        output_path=output,
        owl_image_path=owl,
        program_code="MSCSIA",
    )
    print(f"✓ Wrote {result}")
